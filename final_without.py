import Data
import Function
import Neighborhood
import time
import copy
import random
import Neighborhood11
import Neighborhood10
import Neighborhood_drone
import glob
import os
import openpyxl
import csv
import numpy as np
import math
import sys
import json
global LOOP
global tabu_tenure
global best_sol
global best_fitness
global Tabu_Structure
global current_neighborhood
global LOOP_IMPROVED
global SET_LAST_10
global BEST

# Set up chỉ số -------------------------------------------------------------------
ITE = 1
epsilon = (-1) * 0.00001
# 15:   120,    20:    150
# BREAKLOOP = Data.number_of_cities * 8
LOOP_IMPROVED = 0
SET_LAST_10 = [] 
BEST = []
number_of_cities = int(os.getenv('NUMBER_OF_CITIES')) 
delta = 0.3
alpha = [0.5, 0.3, 0.1]
data_set = str(os.getenv('DATA_SET'))
SEGMENT = 3
solution_pack_len = int(os.getenv('SOLUTION_PACK_LEN'))
similarity = float(os.getenv('SIMILARITY'))
theta = 2
TIME_LIMIT = 18000
def roulette_wheel_selection(population, fitness_scores):
    total_fitness = sum(fitness_scores)
    probabilities = [score / total_fitness for score in fitness_scores]
    selected_index = np.random.choice(len(population), p=probabilities)
    return population[selected_index]

def Tabu_search(init_solution, tabu_tenure, CC, first_time,Data1, index_consider_elite_set, solution_pack):

    current_fitness, current_truck_time, current_sum_fitness = Function.fitness(init_solution)
    best_sol = init_solution
    best_fitness = current_fitness
    sol_chosen_to_break = init_solution
    fit_of_sol_chosen_to_break = current_fitness
    
    lennn = [0] * 6
    lenght_i = [0] * 6
    i = 0
    
    Result_print = []
    # LOOP = BREAKLOOP * AA
    # print(Data.standard_deviation)
    global current_neighborhood
    global LOOP_IMPROVED
    LOOP_IMPROVED = 0
    global use_optimize_truck_route
    use_optimize_truck_route = False
    
    # LOOP = min(int(Data.number_of_cities*math.log10(Data.number_of_cities)), 100)

    # BREAKLOOP = Data.number_of_cities
    END_SEGMENT =  int(Data.number_of_cities/math.log10(Data.number_of_cities)) * theta
    
    T = 0
    nei_set = [0, 1, 2, 3]
    weight = [1/len(nei_set)]*len(nei_set)
    current_sol = init_solution
    data_to_write = {}
    while T < SEGMENT:
        tabu_tenure = tabu_tenure1 = tabu_tenure3 = tabu_tenure2 = random.uniform(2*math.log(Data.number_of_cities), Data.number_of_cities)
        Tabu_Structure = [(tabu_tenure +1) * (-1)] * Data.number_of_cities
        Tabu_Structure1 = [(tabu_tenure +1) * (-1)] * Data.number_of_cities
        Tabu_Structure2 = [(tabu_tenure +1) * (-1)] * Data.number_of_cities
        Tabu_Structure3 = [(tabu_tenure +1) * (-1)] * Data.number_of_cities
        factor = delta #0.3 0.6
        score = [0]*len(nei_set)
        used = [0]*len(nei_set)
        prev_f = best_fitness
        prev_fitness = current_fitness
        
        LOOP_IMPROVED = 0
        lennn = [0] * 6
        lenght_i = [0] * 6
        i = 0
        while i < END_SEGMENT:
            current_neighborhood = []
            choose = roulette_wheel_selection(nei_set, weight)
            if choose == 0:
                current_neighborhood1, solution_pack = Neighborhood.Neighborhood_combine_truck_and_drone_neighborhood_with_tabu_list_with_package(name_of_truck_neiborhood=Neighborhood10.Neighborhood_one_opt_standard, solution=current_sol, number_of_potial_solution=CC, number_of_loop_drone=2, tabu_list=Tabu_Structure, tabu_tenure=tabu_tenure,  index_of_loop=lenght_i[1], best_fitness=best_fitness, kind_of_tabu_structure=1, need_truck_time=False, solution_pack=solution_pack, solution_pack_len=solution_pack_len, use_solution_pack=first_time, index_consider_elite_set=index_consider_elite_set)
                current_neighborhood.append([1, current_neighborhood1])
            elif choose == 2:
                current_neighborhood5, solution_pack = Neighborhood.Neighborhood_combine_truck_and_drone_neighborhood_with_tabu_list_with_package(name_of_truck_neiborhood=Neighborhood11.Neighborhood_two_opt_tue, solution=current_sol, number_of_potial_solution=CC, number_of_loop_drone=2, tabu_list=Tabu_Structure3, tabu_tenure=tabu_tenure3,  index_of_loop=lenght_i[5], best_fitness=best_fitness, kind_of_tabu_structure=5, need_truck_time=False, solution_pack=solution_pack, solution_pack_len=solution_pack_len, use_solution_pack=first_time, index_consider_elite_set=index_consider_elite_set)
                current_neighborhood.append([5, current_neighborhood5])
            elif choose == 3: 
                current_neighborhood4, solution_pack = Neighborhood.Neighborhood_combine_truck_and_drone_neighborhood_with_tabu_list_with_package(name_of_truck_neiborhood=Neighborhood11.Neighborhood_move_2_1, solution=current_sol, number_of_potial_solution=CC, number_of_loop_drone=2, tabu_list=Tabu_Structure2, tabu_tenure=tabu_tenure2,  index_of_loop=lenght_i[4], best_fitness=best_fitness, kind_of_tabu_structure=4, need_truck_time=False, solution_pack=solution_pack, solution_pack_len=solution_pack_len, use_solution_pack=first_time, index_consider_elite_set=index_consider_elite_set)
                current_neighborhood.append([4, current_neighborhood4])
            else:
                current_neighborhood3, solution_pack = Neighborhood.Neighborhood_combine_truck_and_drone_neighborhood_with_tabu_list_with_package(name_of_truck_neiborhood=Neighborhood11.Neighborhood_move_1_1_ver2, solution=current_sol, number_of_potial_solution=CC, number_of_loop_drone=2, tabu_list=Tabu_Structure1, tabu_tenure=tabu_tenure1,  index_of_loop=lenght_i[3], best_fitness=best_fitness, kind_of_tabu_structure=3, need_truck_time=False, solution_pack=solution_pack, solution_pack_len=solution_pack_len, use_solution_pack=first_time, index_consider_elite_set=index_consider_elite_set)
                current_neighborhood.append([3, current_neighborhood3])

            flag = False
            index = [0] * len(current_neighborhood)
            min_nei = [100000] * len(current_neighborhood)
            min_sum = [1000000000] * len(current_neighborhood)
            # print(current_neighborhood)
            for j in range(len(current_neighborhood)):
                if current_neighborhood[j][0] in [1, 2]:
                    for k in range(len(current_neighborhood[j][1])):
                        cfnode = current_neighborhood[j][1][k][1][0]
                        if cfnode - best_fitness < epsilon:
                            min_nei[j] = cfnode
                            index[j] = k
                            best_fitness = cfnode
                            best_sol = current_neighborhood[j][1][k][0]
                            LOOP_IMPROVED = i
                            flag = True

                        elif cfnode - min_nei[j] < epsilon and Tabu_Structure[current_neighborhood[j][1][k][2]] + tabu_tenure <= lenght_i[1]:
                            min_nei[j] = cfnode
                            index[j] = k
                            min_sum[j] = current_neighborhood[j][1][k][1][2]

                        elif min_nei[j] - epsilon > cfnode and Tabu_Structure[current_neighborhood[j][1][k][2]] + tabu_tenure <= lenght_i[1]:
                            if min_sum[j] > current_neighborhood[j][1][k][1][2]:
                                min_nei[j] = cfnode
                                index[j] = k
                                min_sum[j] = current_neighborhood[j][1][k][1][2]
                elif current_neighborhood[j][0] == 3:
                    for k in range(len(current_neighborhood[j][1])):    
                        cfnode = current_neighborhood[j][1][k][1][0]
                        if cfnode - best_fitness < epsilon:
                            min_nei[j] = cfnode
                            index[j] = k
                            best_fitness = cfnode
                            best_sol = current_neighborhood[j][1][k][0]
                            LOOP_IMPROVED = i
                            flag = True

                        elif cfnode - min_nei[j] < epsilon and Tabu_Structure1[current_neighborhood[j][1][k][2][0]] + tabu_tenure1 <= lenght_i[3] or Tabu_Structure1[current_neighborhood[j][1][k][2][1]] + tabu_tenure1 <= lenght_i[3]:
                            min_nei[j] = cfnode
                            index[j] = k
                            min_sum[j] = current_neighborhood[j][1][k][1][2]

                        elif cfnode < min_nei[j] - epsilon and Tabu_Structure1[current_neighborhood[j][1][k][2][0]] + tabu_tenure1 <= lenght_i[3] or Tabu_Structure1[current_neighborhood[j][1][k][2][1]] + tabu_tenure1 <= lenght_i[3]:
                            if min_sum[j] > current_neighborhood[j][1][k][1][2]:
                                min_nei[j] = cfnode
                                index[j] = k
                                min_sum[j] = current_neighborhood[j][1][k][1][2]
                elif current_neighborhood[j][0] == 4:
                    for k in range(len(current_neighborhood[j][1])):
                        cfnode = current_neighborhood[j][1][k][1][0]
                        if cfnode - best_fitness < epsilon:
                            min_nei[j] = cfnode
                            index[j] = k
                            best_fitness = cfnode
                            best_sol = current_neighborhood[j][1][k][0]
                            LOOP_IMPROVED = i
                            flag = True

                        elif cfnode - min_nei[j] < epsilon and Tabu_Structure2[current_neighborhood[j][1][k][2][0]] + tabu_tenure2 <= lenght_i[4] or Tabu_Structure2[current_neighborhood[j][1][k][2][1]] + tabu_tenure2 <= lenght_i[4] or Tabu_Structure2[current_neighborhood[j][1][k][2][2]] + tabu_tenure2 <= lenght_i[4]:
                            min_nei[j] = cfnode
                            index[j] = k
                            min_sum[j] = current_neighborhood[j][1][k][1][2]
                            
                        elif cfnode < min_nei[j] - epsilon and Tabu_Structure2[current_neighborhood[j][1][k][2][0]] + tabu_tenure2 <= lenght_i[4] or Tabu_Structure2[current_neighborhood[j][1][k][2][1]] + tabu_tenure2 <= lenght_i[4] or Tabu_Structure2[current_neighborhood[j][1][k][2][2]] + tabu_tenure2 <= lenght_i[4]:
                            if min_sum[j] > current_neighborhood[j][1][k][1][2]:
                                min_nei[j] = cfnode
                                index[j] = k
                                min_sum[j] = current_neighborhood[j][1][k][1][2]
                elif current_neighborhood[j][0] == 5:
                    for k in range(len(current_neighborhood[j][1])):    
                        cfnode = current_neighborhood[j][1][k][1][0]
                        if cfnode - best_fitness < epsilon:
                            min_nei[j] = cfnode
                            index[j] = k
                            best_fitness = cfnode
                            best_sol = current_neighborhood[j][1][k][0]
                            LOOP_IMPROVED = i
                            flag = True

                        elif cfnode - min_nei[j] < epsilon and Tabu_Structure3[current_neighborhood[j][1][k][2][0]] + tabu_tenure3 <= lenght_i[5] or Tabu_Structure3[current_neighborhood[j][1][k][2][1]] + tabu_tenure3 <= lenght_i[5]:
                            min_nei[j] = cfnode
                            index[j] = k
                            min_sum[j] = current_neighborhood[j][1][k][1][2]

                        elif cfnode < min_nei[j] - epsilon and Tabu_Structure3[current_neighborhood[j][1][k][2][0]] + tabu_tenure3 <= lenght_i[5] or Tabu_Structure3[current_neighborhood[j][1][k][2][1]] + tabu_tenure3 <= lenght_i[5]:
                            if min_sum[j] > current_neighborhood[j][1][k][1][2]:
                                min_nei[j] = cfnode
                                index[j] = k
                                min_sum[j] = current_neighborhood[j][1][k][1][2]
                else:
                    for k in range(len(current_neighborhood[j][1])):
                        cfnode = current_neighborhood[j][1][k][1][0]
                        if cfnode - best_fitness < epsilon:
                            min_nei[j] = cfnode
                            index[j] = k
                            best_fitness = cfnode
                            best_sol = current_neighborhood[j][1][k][0]
                            LOOP_IMPROVED = i
                            flag = True
                            
                        elif cfnode - min_nei[j] < epsilon:
                            min_nei[j] = cfnode
                            index[j] = k
                            min_sum[j] = current_neighborhood[j][1][k][1][2]
                            
                        elif cfnode < min_nei[j] - epsilon:
                            if min_sum[j] > current_neighborhood[j][1][k][1][2]:
                                min_nei[j] = cfnode
                                index[j] = k
                                min_sum[j] = current_neighborhood[j][1][k][1][2]
            index_best_nei = 0
            best_fit_in_cur_loop = min_nei[0]
            
            # for j in range(len(min_nei)):
            #     print(min_nei[j])
            #     print(current_neighborhood[j][1][index[j]][0])
            #     print("-------")
            
            for j in range(1, len(min_nei)):
                if min_nei[j] < best_fit_in_cur_loop:
                    index_best_nei = j
                    best_fit_in_cur_loop = min_nei[j]
            
            if current_neighborhood[index_best_nei][0] in [1, 2]:
                lenght_i[1] += 1
            
            if current_neighborhood[index_best_nei][0] == 3:
                lenght_i[3] += 1
                
            if current_neighborhood[index_best_nei][0] == 4:
                lenght_i[4] += 1
                
            if current_neighborhood[index_best_nei][0] == 5:
                lenght_i[5] += 1
                
            # print(current_neighborhood[index_best_nei][0])
            # print(len(current_neighborhood[index_best_nei][1]))
            # print(current_neighborhood[index_best_nei][1])
            # print(lenght_i[1], " then ", Tabu_Structure)
            # print(lenght_i[3], " then ", Tabu_Structure1)
            # print(lenght_i[4], " then ", Tabu_Structure2)
            # print(lenght_i[5], " then ", Tabu_Structure3)

            if len(current_neighborhood[index_best_nei][1]) == 0:
                # print("hahhaa")
                continue
                
            # print(index[index_best_nei])
            current_sol = current_neighborhood[index_best_nei][1][index[index_best_nei]][0]
            current_fitness = current_neighborhood[index_best_nei][1][index[index_best_nei]][1][0]
            current_truck_time = current_neighborhood[index_best_nei][1][index[index_best_nei]][1][1]
            current_sum_fitness = current_neighborhood[index_best_nei][1][index[index_best_nei]][1][2]
            print(current_fitness, current_sol)
            Data1.append(current_fitness)
            Data1.append(current_sol)
            # SET_LAST_10.append([current_sol, [current_fitness, current_truck_time]])
            # if len(SET_LAST_10) > 10:
            #     SET_LAST_10.pop(0)
            
            if current_neighborhood[index_best_nei][0] in [1, 2]:
                Tabu_Structure[current_neighborhood[index_best_nei][1][index[index_best_nei]][2]] = lenght_i[1] -1
                lennn[current_neighborhood[index_best_nei][0]] += 1
            
            if current_neighborhood[index_best_nei][0] == 3:
                Tabu_Structure1[current_neighborhood[index_best_nei][1][index[index_best_nei]][2][0]] = lenght_i[3] - 1 
                Tabu_Structure1[current_neighborhood[index_best_nei][1][index[index_best_nei]][2][1]] = lenght_i[3] - 1
                lennn[current_neighborhood[index_best_nei][0]] += 1
                
            if current_neighborhood[index_best_nei][0] == 4:
                Tabu_Structure2[current_neighborhood[index_best_nei][1][index[index_best_nei]][2][0]] = lenght_i[4] - 1
                Tabu_Structure2[current_neighborhood[index_best_nei][1][index[index_best_nei]][2][1]] = lenght_i[4] - 1
                Tabu_Structure2[current_neighborhood[index_best_nei][1][index[index_best_nei]][2][2]] = lenght_i[4] - 1
                lennn[current_neighborhood[index_best_nei][0]] += 1
                
            if current_neighborhood[index_best_nei][0] == 5:
                Tabu_Structure3[current_neighborhood[index_best_nei][1][index[index_best_nei]][2][0]] = lenght_i[5] - 1
                Tabu_Structure3[current_neighborhood[index_best_nei][1][index[index_best_nei]][2][1]] = lenght_i[5] - 1
                lennn[current_neighborhood[index_best_nei][0]] += 1
                
            if fit_of_sol_chosen_to_break > current_fitness:
                sol_chosen_to_break = current_sol
                fit_of_sol_chosen_to_break = current_fitness
                LOOP_IMPROVED = i
                
            

            if current_neighborhood[index_best_nei][0] in [1, 2]:
                temp = [current_neighborhood[index_best_nei][0], current_fitness, current_neighborhood[index_best_nei][1][index[index_best_nei]][2], -1, current_sol, Tabu_Structure, Tabu_Structure1]
            elif current_neighborhood[index_best_nei][0] in [3]:
                temp = [current_neighborhood[index_best_nei][0], current_fitness, current_neighborhood[index_best_nei][1][index[index_best_nei]][2][0], current_neighborhood[index_best_nei][1][index[index_best_nei]][2][1], current_sol, Tabu_Structure, Tabu_Structure1]
            else:
                temp = [current_neighborhood[index_best_nei][0], current_fitness, -1, -1, current_sol]
            Data1.append(temp)

            used[choose] += 1
            if flag == True:
                score[choose] += alpha[0]
            elif current_fitness - prev_fitness < epsilon:
                score[choose] += alpha[1]
            else:
                score[choose] += alpha[2]

            for j in range(len(nei_set)):
                if used[j] == 0:
                    continue
                else:
                    weight[j] = (1 - factor)*weight[j] + factor*score[j]/used[j]
            if flag == True:
                i = 0
            else:
                i += 1
        print("-------",T,"--------")
        print(best_fitness)
        print(T, best_sol, "\n", best_fitness)
        print(used, score, sum(used))

        if best_fitness - prev_f < epsilon:
            T = 0
        else: 
            T += 1
        
    return best_sol, best_fitness, Result_print, solution_pack
    
def Tabu_search_for_CVRP(CC):
    with open('Random_'+str(data_set)+'_'+str(number_of_cities)+'_'+str(solution_pack_len)+'_'+str(similarity)+'_withoutdiv.json', 'r') as file:
        lines = file.readlines()
        last_line = lines[-1]
        data = json.loads(last_line)  # Parse the last line as JSON
        done = data["Done"]
    if done == True:
        return data["best_sol"], data["best_fitness"], data["runtime"]
    else:
        best_sol = data["best_sol"]
        best_fitness = float(data["best_fitness"])
        solution_pack = data["solution_pack"]
        
        runtime = data["runtime"]
        Data1 = []

        for pi in range(solution_pack_len):
            if pi < len(solution_pack):
                # current_neighborhood5 = Neighborhood.swap_two_array(solution_pack[pi][0])
                # best_sol_in_brnei = current_neighborhood5[0][0]
                # best_fitness_in_brnei = current_neighborhood5[0][1][0]
                # for i in range(1, len(current_neighborhood5)):
                #     cfnode = current_neighborhood5[i][1][0]
                #     if cfnode - best_fitness_in_brnei < epsilon:
                #         best_sol_in_brnei = current_neighborhood5[i][0]
                #         best_fitness_in_brnei = cfnode
                best_sol_in_brnei = solution_pack[pi][0]
                best_sol1, best_fitness1, result_print1, solution_pack = Tabu_search(init_solution=best_sol_in_brnei, tabu_tenure=Data.number_of_cities-1, CC=CC, first_time=False,Data1=Data1, index_consider_elite_set=pi+1, solution_pack=solution_pack)
                # print("-----------------", pi, "------------------------")
                # print(best_sol1)
                # print(best_fitness1)
                if best_fitness1 - best_fitness < epsilon:
                    best_sol = best_sol1
                    best_fitness = best_fitness1

        return best_sol, best_fitness, runtime

# Thư mục chứa các file .txt
folder_path = "test_data/data_demand_random/"+str(number_of_cities)

txt_files = glob.glob(os.path.join(folder_path, data_set))
workbook = openpyxl.Workbook()
sheet = workbook.active
row = 1
column = 1

# Ghi tên file .txt vào cột đầu tiên
for txt_file in txt_files:
    sheet.cell(row=row, column=column, value=os.path.basename(txt_file))
    row += 1
# Đặt lại dòng và cột cho việc ghi kết quả
row = 1
for txt_file in txt_files:
    column = 2
    with open(txt_file, 'r') as file:
        Data.read_data_random(txt_file)
        result = []
        run_time = []
        avg = 0
        avg_run_time = 0
        best_csv_fitness = 1000000
        for i in range(ITE):
            BEST = []
            print("------------------------",i,"------------------------")
            start_time = time.time()
            best_sol, best_fitness, runtime = Tabu_search_for_CVRP(1)
            print("---------- RESULT ----------")
            print(best_sol)
            print(best_fitness)
            result.append(best_fitness)
            # print(Function.Check_if_feasible(best_sol))
            end_time = time.time()
            run = end_time - start_time + runtime
            run_time.append(run)
            avg_run_time += run/ITE
            sheet.cell(row=row, column=column, value=best_fitness)

            column += 1
            if best_csv_fitness > best_fitness:
                best_csv_sol = best_sol
                best_csv_fitness = best_fitness
            if i == ITE - 1:
                sheet.cell(row=row, column=column, value=avg_run_time)
                sheet.cell(row=row, column=column+1, value=str(best_csv_sol))    
        # Tăng dòng cho lần chạy tiếp theo
        row += 1
    workbook.save(f"Random_{data_set}_{number_of_cities}_{solution_pack_len}_{similarity}_withoutdiv.xlsx")
        # log_file.close()

workbook.close()